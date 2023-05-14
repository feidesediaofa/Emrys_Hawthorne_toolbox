import os
import os.path
import tkinter as tk
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText as Text
from tkinter import simpledialog, filedialog, messagebox
from openpyxl import Workbook
import pyperclip
import datetime
import pickle
import threading
import queue
import time
import sys
from infi.systray import SysTrayIcon 
import traceback
import win32event
import win32api
import winerror

if getattr(sys, 'frozen', False):
    os.environ['TCL_LIBRARY'] = os.path.join(sys._MEIPASS, "tcl")
    os.environ['TK_LIBRARY'] = os.path.join(sys._MEIPASS, "tk")
    os.environ['NUMPY_MKL_DLL_PATH'] = sys._MEIPASS


class ClipboardHistoryApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Clipboard History")
        self.history_file = "clipboard_history.pkl"

        self.load_history()

        # Create treeview with scrollbars
        self.tree_frame = tk.Frame(self)
        self.tree = ttk.Treeview(self.tree_frame, columns=("Date", "Content", "Total Copies", "Fav", "Name", "Note"), show="headings")
        self.vsb = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        self.hsb = ttk.Scrollbar(self.tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=self.vsb.set, xscrollcommand=self.hsb.set)

        self.filtered_history = self.history

        self.tree.heading("Date", text="Date")
        self.tree.heading("Content", text="Content")
        self.tree.heading("Total Copies", text="Total Copies")
        self.tree.heading("Fav", text="Fav")
        self.tree.heading("Name", text="Name")
        self.tree.heading("Note", text="Note")
        self.tree.bind("<Double-1>", self.on_item_double_click)
        self.fill_tree()



        self.tree.grid(row=0, column=0, sticky='nsew')
        self.vsb.grid(row=0, column=1, sticky='ns')
        self.hsb.grid(row=1, column=0, sticky='ew')
        self.tree_frame.rowconfigure(0, weight=1)
        self.tree_frame.columnconfigure(0, weight=1)
        self.tree_frame.pack(fill=tk.BOTH, expand=True)

        self.search_frame = tk.Frame(self)
        self.search_box = ttk.Entry(self.search_frame)
        self.search_box.pack(side=tk.LEFT, padx=5)
        self.search_button = ttk.Button(self.search_frame, text="Search", command=self.on_search_button_click)
        self.search_button.pack(side=tk.LEFT, padx=5)
        self.search_frame.pack(side=tk.TOP, fill=tk.X, pady=5)        

        self.copy_button = ttk.Button(self, text="Copy", command=self.on_copy_button_click)
        self.copy_button.pack(side=tk.LEFT)

        self.delete_button = ttk.Button(self, text="Delete", command=self.on_delete_button_click)
        self.delete_button.pack(side=tk.LEFT)

        self.favorite_button = ttk.Button(self, text="Favorite", command=self.on_favorite_button_click)
        self.favorite_button.pack(side=tk.LEFT)

        self.filter_button = ttk.Button(self, text="Filter Favorites", command=self.on_filter_button_click)
        self.filter_button.pack(side=tk.LEFT)
        
        export_button_frame = ttk.Frame(self)
        export_button = ttk.Button(export_button_frame, text="Export Selected Rows", command=self.export_selected_rows_to_excel)
        export_button.pack(padx=5, pady=5)
        export_button_frame.pack(side=tk.LEFT, padx=5)

        self.clipboard_last = ""
        self.show_favorites_only = False
        
        # Start the clipboard monitor daemon thread
        daemon_thread = threading.Thread(target=self.start_clipboard_monitor, daemon=True)
        daemon_thread.start()

        self.clipboard_queue = queue.Queue()
        self.process_clipboard_queue()
        
        # Create system tray icon
        if sys.platform == "win32":
            self.icon = SysTrayIcon(r"E:\\clipboard_history\\clipboard_history.ico", "Clipboard History",
                                    (("Open", None, self.show_window),
                                     ("Exit", None, self.on_exit)))
        else:
            # macOS implementation with rumps
            pass

        self.protocol("WM_DELETE_WINDOW", self.minimize_to_tray)

        
    def on_item_double_click(self, event):
        item_id = self.tree.selection()[0]
        col = self.tree.identify_column(event.x)

        if col == "#5":  # Name column
            self.update_name(item_id)
        elif col == "#6":  # Note column
            self.update_note(item_id)
        elif col == "#4":  # Fav column
            self.on_favorite_button_click(ignore_update=True)  # Add ignore_update=True
        else:
            _, content, _, _, _, _ = self.tree.item(item_id, "values")

            content_window = tk.Toplevel(self)
            content_window.title("Content")
            text_widget = Text(content_window, wrap=tk.WORD)
            text_widget.insert(tk.END, content)
            text_widget.pack(expand=True, fill=tk.BOTH)



    def update_name(self, item_id):
        _, content, _, _, old_name, _ = self.tree.item(item_id)['values']
        new_name = simpledialog.askstring("Name", "Enter a new name:", initialvalue=old_name)

        if new_name is not None:
            self.tree.set(item_id, "Name", new_name)
            self.history[content]["name"] = new_name
            self.save_history()
            
    def minimize_to_tray(self):
        self.icon.visible = True
        self.withdraw()
        threading.Thread(target=self.icon.start).start()
    
    def show_window(self, systray):
        self.deiconify()
    
    def on_exit(self, systray):
        def delayed_shutdown():
            systray.shutdown()
            self.destroy()
    
        # Use a Timer to call the shutdown and destroy methods in a new thread after a short delay
        timer = threading.Timer(0.1, delayed_shutdown)
        timer.start()
   

    def update_note(self, item_id):
        _, content, _, _, _, old_note = self.tree.item(item_id)['values']
        new_note = simpledialog.askstring("Note", "Enter a new note:", initialvalue=old_note)

        if new_note is not None:
            self.tree.set(item_id, "Note", new_note)
            self.history[content]["note"] = new_note
            self.save_history()
        

    def load_history(self):
        if os.path.exists(self.history_file):
            try:
                with open(self.history_file, "rb") as f:
                    self.history = pickle.load(f)
            except Exception as e:
                print(f"Error loading history file: {e}")
                traceback.print_exc()  # 添加这一行以打印堆栈跟踪
                self.history = {}
        else:
            self.history = {}



    def save_history(self):
        try:
            with open(self.history_file, "wb") as f:
                pickle.dump(self.history, f)
        except Exception as e:
            print(f"Error saving history file: {e}")
            
    def fill_tree(self):
        # Clear the treeview before filling it
        self.tree.delete(*self.tree.get_children())

        for content, data in self.filtered_history.items():
            fav = data.get("fav", "No")
            name = data.get("name", "")
            note = data.get("note", "")
            self.tree.insert('', 'end', content, values=(data['Date'], content, data['Total Copies'], fav, name, note))

    def check_clipboard(self):
        content = pyperclip.paste()
        if content != self.clipboard_last and content not in self.history:
            self.clipboard_last = content
            date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            self.history[content] = {'date': date, 'copies': 0}
            self.tree.insert('', 'end', content, values=(date, content, 0))
            self.save_history()

        self.after(500, self.check_clipboard)


    # ... (same as before: load_history, save_history, fill_tree, check_clipboard, on_item_double_click) ...

    def on_copy_button_click(self):
        item_id = self.tree.selection()
        if not item_id:
            return

        item_id = item_id[0]
        _, content, copies, _, _, _ = self.tree.item(item_id)['values']
        pyperclip.copy(content)
        updated_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated_copies = int(copies) + 1
        self.tree.set(item_id, "Date", updated_date)
        self.tree.set(item_id, "Total Copies", updated_copies)

        # Update the 'copies' value in the self.history dictionary
        self.history[content]["date"] = updated_date
        self.history[content]["copies"] = updated_copies

        self.save_history()


    def on_delete_button_click(self):
        item_id = self.tree.selection()
        if not item_id:
            return
        
        # Add confirmation messagebox
        confirmed = messagebox.askyesno("Delete", "Are you sure you want to delete this item?")
        if not confirmed:
            return

        item_id = item_id[0]
        _, content, _, _, _, _ = self.tree.item(item_id)['values']
        self.tree.delete(item_id)
        if content in self.history:
            del self.history[content]
        self.save_history()
    def start_clipboard_monitor(self):
        while True:
            try:
                clipboard_current = pyperclip.paste()
                if clipboard_current != self.clipboard_last and clipboard_current.strip() != "":
                    self.clipboard_last = clipboard_current
                    self.clipboard_queue.put(clipboard_current)
            except pyperclip.exceptions.PyperclipException:
                pass
            time.sleep(1)
            
    def add_history_item(self, content, date_time=None, total_copies=1, fav=False, name="", note=""):
        if date_time is None:
            date_time = datetime.datetime.now()
        item = {
            "Date": date_time.strftime("%Y-%m-%d %H:%M:%S"),
            "Content": content,
            "Total Copies": total_copies,
            "Fav": fav,
            "Name": name,
            "Note": note,
        }
        
        self.history[content] = item  # 添加这一行以更新 self.history 字典
        
        self.tree.insert("", 0, values=(item["Date"], item["Content"], item["Total Copies"], item["Fav"], item["Name"], item["Note"]))
                
    def process_clipboard_queue(self):
        try:
            while True:
                clipboard_content = self.clipboard_queue.get_nowait()
                item_exists = False
                for row_id in self.tree.get_children():
                    row_values = self.tree.item(row_id, "values")
                    if len(row_values) >= 2 and row_values[1] == clipboard_content:
                        item_exists = True
                        break

                if not item_exists:
                    self.add_history_item(clipboard_content)
                    self.save_history()

        except queue.Empty:
            pass

        self.after(1000, self.process_clipboard_queue)

    def on_search_button_click(self):
        search_term = self.search_box.get().strip()
        if not search_term:
            self.filtered_history = self.history
        else:
            self.filtered_history = {content: data for content, data in self.history.items() if hasattr(content, 'lower') and search_term.lower() in content.lower()}
            
        self.fill_tree()


        
    def export_selected_rows_to_excel(self):
            selected_items = self.tree.selection()
            if not selected_items:
                tk.messagebox.showinfo("No Selection", "Please select at least one row to export.")
                return
    
            # Create a new workbook and worksheet
            workbook = Workbook()
            sheet = workbook.active
    
            # Write header row
            headers = ["Timestamp", "Content", "Name", "Note"]
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)
    
            # Write data rows
            for idx, item_id in enumerate(selected_items, start=2):
                timestamp, content, _, _, name, note = self.tree.item(item_id)["values"]
    
                sheet.cell(row=idx, column=1, value=timestamp)
                sheet.cell(row=idx, column=2, value=content)
                sheet.cell(row=idx, column=3, value=name)
                sheet.cell(row=idx, column=4, value=note)
    
            # Ask user for save location
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    
            if file_path:
                workbook.save(file_path)
                tk.messagebox.showinfo("Export Successful", f"Data exported to {file_path}")

    def on_favorite_button_click(self, ignore_update=False):  # Add ignore_update parameter with default value False
        item_id = self.tree.selection()
        if not item_id:
            return
        item_id = item_id[0]
        _, content, _, fav, _, _ = self.tree.item(item_id)['values']

        if fav == "Yes":
            new_fav = "No"
        else:
            new_fav = "Yes"

        self.tree.set(item_id, "Fav", new_fav)
        self.history[content]["fav"] = new_fav

        # Add this condition to decide whether to update the name and note when the favorite button is clicked
        if not ignore_update:
            name = simpledialog.askstring("Name", "Enter a name:")
            note = simpledialog.askstring("Note", "Enter a note:")

            self.tree.set(item_id, "Name", name)
            self.tree.set(item_id, "Note", note)

            self.history[content]["name"] = name
            self.history[content]["note"] = note

        self.save_history()


    def on_filter_button_click(self):
        self.show_favorites_only = not self.show_favorites_only

        if self.show_favorites_only:
            self.filter_button.config(text="Show All")
            # Update the filtered_history with favorites only
            self.filtered_history = {k: v for k, v in self.history.items() if v.get("fav") == "Yes"}
        else:
            self.filter_button.config(text="Filter Favorites")
            self.filtered_history = self.history  # Show all items when not filtering

        # 清除树中的所有项目
        for item in self.tree.get_children():
            self.tree.delete(item)

        # 根据当前过滤设置重新填充树
        for content, data in self.filtered_history.items():  # Use self.filtered_history instead of self.history
            fav = data.get("fav", "No")

            self.tree.insert('', 'end', content, values=(
                data['Date'], content, data['Total Copies'], fav,  # Change data['Copies'] to data['Total Copies']
                data.get('name', ''), data.get('note', '')
            ))

# 创建互斥对象
mutex = win32event.CreateMutex(None, False, "clipboard_history_mutex")

# 检查是否已经有一个实例在运行
if win32api.GetLastError() == winerror.ERROR_ALREADY_EXISTS:
    # 显示一个错误对话框
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    messagebox.showerror("你已经打开一个了！", "现在，点确定，去右下角系统托盘右键然后单击open使用上一次的！")
    sys.exit()


if __name__ == "__main__":
    app = ClipboardHistoryApp()
    app.mainloop()
